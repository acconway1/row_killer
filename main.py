import pandas as pd
import os
import openpyxl

def delete_blank_rows(file_path):
    # Determine the new file name with "_no_blanks.xlsx" appended
    base, ext = os.path.splitext(file_path)
    new_file_path = base + "_no_blanks" + ext

    # Load the workbook and process each sheet
    wb = openpyxl.load_workbook(file_path)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_to_delete = []

        # Find rows where all elements are None or empty
        for row in ws.iter_rows():
            if all(cell.value is None or cell.value == "" for cell in row):
                rows_to_delete.append(row[0].row)

        # Delete rows in reverse order to avoid shifting issues
        for row in reversed(rows_to_delete):
            ws.delete_rows(row, 1)
    
    # Save the workbook to a new file
    wb.save(new_file_path)

if __name__ == "__main__":
    file_path = input("Enter the path to your Excel file: ")
    delete_blank_rows(file_path)
    print("Blank rows deleted successfully. New file created without blank rows.")
